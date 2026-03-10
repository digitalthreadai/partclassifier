"""
Part Classification Agent -- Streamlit UI
=========================================
Run with:  streamlit run app.py
"""

import asyncio
import os
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv

from src.llm_client import PROVIDER_PRESETS

# -- Paths -----------------------------------------------------------------
BASE_DIR = Path(__file__).parent
ENV_PATH = BASE_DIR / ".env"
DEFAULT_INPUT_DIR = str(BASE_DIR / "input")
OUTPUT_DIR = BASE_DIR / "output"

# -- Provider display names (derived from llm_client.py) -------------------
PROVIDER_DISPLAY = {
    preset["label"]: key for key, preset in PROVIDER_PRESETS.items()
}

# -- Provider emoji icons (no external URLs -- always works) ---------------
PROVIDER_EMOJI = {
    "groq": "&#9889;",       # lightning bolt
    "openai": "&#9679;",     # circle (GPT style)
    "anthropic": "&#10022;", # star
    "ollama": "&#128038;",   # bird
}


# -- Custom CSS for modern design -----------------------------------------
def inject_custom_css(dark_mode: bool = True):
    # Dark theme variables
    tv = {
        "bg": "#0f172a", "bg_secondary": "#1e293b", "bg_card": "#1e293b",
        "border": "rgba(148, 163, 184, 0.15)", "border_hover": "rgba(99, 102, 241, 0.4)",
        "text": "#f1f5f9", "text_secondary": "#94a3b8", "text_muted": "#64748b",
        "stat_bg": "#0f172a", "table_header_bg": "#0f172a",
        "table_hover": "rgba(255,255,255,0.03)", "card_shadow": "0 1px 3px rgba(0,0,0,0.2)",
        "card_shadow_hover": "0 4px 12px rgba(0,0,0,0.3)",
        "icon_blue_bg": "rgba(59,130,246,0.15)", "icon_green_bg": "rgba(22,163,98,0.15)",
        "icon_purple_bg": "rgba(147,51,234,0.15)", "icon_amber_bg": "rgba(217,119,6,0.15)",
    }

    st.markdown(f"""
    <style>
    /* -- Global -- */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    .stApp {{
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        background-color: {tv['bg']} !important;
        color: {tv['text']} !important;
    }}
    .stApp [data-testid="stAppViewContainer"] {{
        background-color: {tv['bg']} !important;
    }}
    .stApp [data-testid="stMain"] {{
        background-color: {tv['bg']} !important;
    }}

    /* -- Hide default Streamlit branding (keep header for sidebar toggle) -- */
    #MainMenu, footer, [data-testid="stDecoration"],
    .stDeployButton, button[kind="header"] {{ display: none !important; visibility: hidden !important; }}
    header[data-testid="stHeader"] {{
        background: transparent !important;
        height: 2rem !important;
    }}

    /* -- Reduce top padding on main area and sidebar -- */
    .stApp [data-testid="stAppViewContainer"] > div:first-child {{
        padding-top: 0 !important;
    }}
    section[data-testid="stSidebar"] > div:first-child {{
        padding-top: 1rem !important;
    }}
    .block-container {{
        padding-top: 1rem !important;
        padding-bottom: 1rem !important;
    }}

    /* -- Sidebar styling (always dark) -- */
    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
    }}
    section[data-testid="stSidebar"] * {{
        color: #e2e8f0 !important;
    }}
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stTextInput label {{
        color: #94a3b8 !important;
        font-size: 0.8rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }}
    section[data-testid="stSidebar"] hr {{
        border-color: rgba(148, 163, 184, 0.2);
    }}
    section[data-testid="stSidebar"] .stAlert {{
        background: rgba(255,255,255,0.05);
        border-radius: 10px;
    }}

    /* -- Hero banner -- */
    .hero-banner {{
        background: linear-gradient(135deg, #1e293b 0%, #334155 50%, #1e293b 100%);
        border-radius: 16px;
        padding: 24px 32px;
        margin-bottom: 16px;
        position: relative;
        overflow: hidden;
        border: 1px solid rgba(148, 163, 184, 0.1);
    }}
    .hero-banner::before {{
        content: '';
        position: absolute;
        top: -50%;
        right: -20%;
        width: 400px;
        height: 400px;
        background: radial-gradient(circle, rgba(99, 102, 241, 0.15) 0%, transparent 70%);
        border-radius: 50%;
    }}
    .hero-banner::after {{
        content: '';
        position: absolute;
        bottom: -30%;
        left: -10%;
        width: 300px;
        height: 300px;
        background: radial-gradient(circle, rgba(16, 185, 129, 0.1) 0%, transparent 70%);
        border-radius: 50%;
    }}
    .hero-title {{
        font-size: 1.5rem;
        font-weight: 800;
        color: #f8fafc;
        margin: 0 0 4px 0;
        letter-spacing: -0.02em;
        position: relative;
        z-index: 1;
    }}
    .hero-subtitle {{
        font-size: 0.88rem;
        color: #94a3b8;
        margin: 0;
        position: relative;
        z-index: 1;
    }}
    .hero-badge {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(99, 102, 241, 0.15);
        border: 1px solid rgba(99, 102, 241, 0.3);
        border-radius: 99px;
        padding: 4px 14px;
        font-size: 0.78rem;
        font-weight: 600;
        color: #a5b4fc;
        margin-top: 10px;
        position: relative;
        z-index: 1;
    }}

    /* -- Card component (themed) -- */
    .ui-card {{
        background: {tv['bg_card']};
        border-radius: 14px;
        border: 1px solid {tv['border']};
        padding: 16px 20px;
        margin-bottom: 12px;
        box-shadow: {tv['card_shadow']};
        transition: box-shadow 0.2s ease;
    }}
    .ui-card:hover {{
        box-shadow: {tv['card_shadow_hover']};
    }}
    .ui-card-header {{
        display: flex;
        align-items: center;
        gap: 10px;
        margin-bottom: 8px;
    }}
    .ui-card-icon {{
        width: 40px;
        height: 40px;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.2rem;
        flex-shrink: 0;
    }}
    .ui-card-icon.blue {{ background: {tv['icon_blue_bg']}; color: #3b82f6; }}
    .ui-card-icon.green {{ background: {tv['icon_green_bg']}; color: #16a34a; }}
    .ui-card-icon.purple {{ background: {tv['icon_purple_bg']}; color: #9333ea; }}
    .ui-card-icon.amber {{ background: {tv['icon_amber_bg']}; color: #d97706; }}
    .ui-card-title {{
        font-size: 1rem;
        font-weight: 700;
        color: {tv['text']};
        margin: 0;
    }}
    .ui-card-desc {{
        font-size: 0.82rem;
        color: {tv['text_secondary']};
        margin: 0;
    }}

    /* -- Status pill -- */
    .status-pill {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        padding: 5px 14px;
        border-radius: 99px;
        font-size: 0.78rem;
        font-weight: 600;
    }}
    .status-pill.active {{
        background: {tv['icon_green_bg']};
        color: #16a34a;
        border: 1px solid rgba(34,197,94,0.3);
    }}
    .status-pill.inactive {{
        background: {tv['icon_amber_bg']};
        color: #d97706;
        border: 1px solid rgba(217,119,6,0.3);
    }}
    .status-pill .dot {{
        width: 7px;
        height: 7px;
        border-radius: 50%;
    }}
    .status-pill.active .dot {{ background: #16a34a; }}
    .status-pill.inactive .dot {{ background: #d97706; }}

    /* -- Stat boxes (themed) -- */
    .stat-row {{
        display: flex;
        gap: 12px;
        margin: 12px 0;
    }}
    .stat-box {{
        flex: 1;
        background: {tv['stat_bg']};
        border-radius: 12px;
        padding: 14px;
        text-align: center;
        border: 1px solid {tv['border']};
    }}
    .stat-value {{
        font-size: 1.8rem;
        font-weight: 800;
        color: {tv['text']};
        line-height: 1;
    }}
    .stat-value.blue {{ color: #3b82f6; }}
    .stat-value.green {{ color: #16a34a; }}
    .stat-value.purple {{ color: #9333ea; }}
    .stat-label {{
        font-size: 0.75rem;
        font-weight: 600;
        color: {tv['text_muted']};
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-top: 6px;
    }}

    /* -- Result card (themed) -- */
    .result-card {{
        background: {tv['bg_card']};
        border-radius: 12px;
        border: 1px solid {tv['border']};
        padding: 20px 24px;
        margin-bottom: 12px;
        display: flex;
        align-items: center;
        gap: 16px;
        transition: all 0.2s ease;
    }}
    .result-card:hover {{
        border-color: {tv['border_hover']};
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.08);
    }}
    .result-icon {{
        width: 44px;
        height: 44px;
        border-radius: 10px;
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.2rem;
        flex-shrink: 0;
        color: white;
    }}
    .result-info {{ flex: 1; }}
    .result-title {{
        font-size: 0.92rem;
        font-weight: 700;
        color: {tv['text']};
        margin: 0;
    }}
    .result-meta {{
        font-size: 0.8rem;
        color: {tv['text_secondary']};
        margin: 4px 0 0 0;
    }}
    .result-badge {{
        display: inline-flex;
        align-items: center;
        gap: 4px;
        background: {tv['icon_green_bg']};
        border: 1px solid rgba(34,197,94,0.3);
        border-radius: 99px;
        padding: 4px 12px;
        font-size: 0.75rem;
        font-weight: 600;
        color: #16a34a;
        flex-shrink: 0;
    }}

    /* -- Attribute table (themed) -- */
    .attr-table {{
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid {tv['border']};
        margin-top: 12px;
    }}
    .attr-table th {{
        background: {tv['table_header_bg']};
        padding: 10px 16px;
        font-size: 0.75rem;
        font-weight: 700;
        color: {tv['text_secondary']};
        text-transform: uppercase;
        letter-spacing: 0.05em;
        text-align: left;
        border-bottom: 1px solid {tv['border']};
    }}
    .attr-table td {{
        padding: 10px 16px;
        font-size: 0.88rem;
        color: {tv['text']};
        border-bottom: 1px solid {tv['border']};
    }}
    .attr-table tr:last-child td {{ border-bottom: none; }}
    .attr-table tr:hover td {{ background: {tv['table_hover']}; }}
    .attr-key {{
        font-weight: 600;
        color: {tv['text_secondary']};
        width: 40%;
    }}

    /* -- Download button row -- */
    .download-row {{
        display: flex;
        gap: 12px;
        flex-wrap: wrap;
        margin: 16px 0;
    }}

    /* -- Section label -- */
    .section-label {{
        font-size: 0.72rem;
        font-weight: 700;
        color: #94a3b8;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 4px;
    }}

    /* -- Compact sidebar widget spacing -- */
    section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {{
        margin-bottom: 0 !important;
        padding-bottom: 0 !important;
    }}
    section[data-testid="stSidebar"] .stSelectbox,
    section[data-testid="stSidebar"] .stTextInput {{
        margin-bottom: 4px !important;
    }}
    section[data-testid="stSidebar"] .stButton {{
        margin-top: 4px !important;
    }}

    /* -- Processing animation -- */
    .processing-card {{
        background: {tv['bg_secondary']};
        border-radius: 14px;
        border: 1px solid {tv['border']};
        padding: 32px;
        text-align: center;
    }}
    .processing-card h3 {{
        color: {tv['text']};
        margin: 12px 0 4px;
    }}
    .processing-card p {{
        color: {tv['text_secondary']};
        font-size: 0.9rem;
    }}

    /* -- Override Streamlit button styles -- */
    .stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
        border: none;
        border-radius: 10px;
        padding: 12px 28px;
        font-weight: 700;
        font-size: 0.95rem;
        letter-spacing: 0.01em;
        box-shadow: 0 4px 14px rgba(99, 102, 241, 0.3);
        transition: all 0.2s ease;
    }}
    .stButton > button[kind="primary"]:hover {{
        box-shadow: 0 6px 20px rgba(99, 102, 241, 0.45);
        transform: translateY(-1px);
    }}

    /* -- Sidebar header -- */
    .sidebar-logo {{
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 0 0 10px;
        border-bottom: 1px solid rgba(148, 163, 184, 0.15);
        margin-bottom: 10px;
    }}
    .sidebar-logo-icon {{
        width: 36px;
        height: 36px;
        border-radius: 9px;
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1rem;
        color: white;
    }}
    .sidebar-logo-text {{
        font-size: 0.85rem;
        font-weight: 700;
        color: #f1f5f9 !important;
        letter-spacing: -0.01em;
    }}
    .sidebar-logo-sub {{
        font-size: 0.68rem;
        color: #64748b !important;
        font-weight: 500;
    }}

    /* -- Provider card in sidebar -- */
    .provider-status {{
        background: rgba(255,255,255,0.06);
        border-radius: 10px;
        padding: 10px 14px;
        border: 1px solid rgba(255,255,255,0.08);
        margin-top: 4px;
    }}
    .provider-status-row {{
        display: flex;
        align-items: center;
        gap: 10px;
    }}
    .provider-status-icon {{
        width: 32px;
        height: 32px;
        border-radius: 8px;
        background: rgba(255,255,255,0.1);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1rem;
        flex-shrink: 0;
    }}
    .provider-status-info {{
        flex: 1;
    }}
    .provider-status-name {{
        font-size: 0.82rem;
        font-weight: 700;
        color: #f1f5f9 !important;
    }}
    .provider-status-model {{
        font-size: 0.72rem;
        color: #94a3b8 !important;
        font-family: 'SF Mono', 'Fira Code', monospace;
    }}

    /* -- Streamlit overrides for cleaner look -- */
    .stProgress > div > div {{ border-radius: 99px; }}
    .stProgress > div > div > div {{
        background: linear-gradient(90deg, #6366f1, #8b5cf6);
        border-radius: 99px;
    }}

    div[data-testid="stExpander"] {{
        border: 1px solid {tv['border']};
        border-radius: 12px;
        overflow: hidden;
    }}
    div[data-testid="stExpander"] summary {{
        font-weight: 600;
        color: {tv['text']};
    }}

    /* -- Streamlit widget theming -- */
    .stApp label, .stApp .stMarkdown p, .stApp .stMarkdown li {{
        color: {tv['text']} !important;
    }}
    .stApp .stTextInput input, .stApp .stSelectbox > div > div {{
        background: {tv['bg_card']} !important;
        color: {tv['text']} !important;
        border-color: {tv['border']} !important;
    }}
    .stApp .stDataFrame {{
        background: {tv['bg_card']};
    }}
    .stApp [data-testid="stExpander"] {{
        background: {tv['bg_card']};
    }}
    .stApp [data-testid="stExpander"] summary span {{
        color: {tv['text']} !important;
    }}
    .stApp [data-testid="stMetricValue"] {{
        color: {tv['text']} !important;
    }}
    .stApp [data-testid="stMetricLabel"] {{
        color: {tv['text_muted']} !important;
    }}
    .stApp [data-testid="stCaptionContainer"] {{
        color: {tv['text_secondary']} !important;
    }}
    .stApp .stAlert {{
        background: rgba(255,255,255,0.05);
    }}
    </style>
    """, unsafe_allow_html=True)


# -- Helpers ---------------------------------------------------------------

def read_env() -> dict[str, str]:
    """Read the .env file into a dict."""
    env = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def write_env(provider: str, api_key: str, model: str):
    """Write LLM config to .env file, preserving other variables."""
    env = read_env()
    env["LLM_PROVIDER"] = provider
    env["LLM_API_KEY"] = api_key
    env["LLM_MODEL"] = model
    env.pop("GROQ_API_KEY", None)

    lines = [f"{k}={v}" for k, v in env.items()]
    ENV_PATH.write_text("\n".join(lines) + "\n")

    os.environ["LLM_PROVIDER"] = provider
    os.environ["LLM_API_KEY"] = api_key
    os.environ["LLM_MODEL"] = model


def get_excel_files(folder: str) -> list[Path]:
    """Return all .xlsx files in a folder."""
    p = Path(folder)
    if not p.exists():
        return []
    return sorted(p.glob("*.xlsx"))


def render_attr_table(attributes: dict) -> str:
    """Render attributes as a styled HTML table."""
    rows = ""
    for k, v in attributes.items():
        rows += f'<tr><td class="attr-key">{k}</td><td>{v}</td></tr>'
    return f'<table class="attr-table"><thead><tr><th>Attribute</th><th>Value</th></tr></thead><tbody>{rows}</tbody></table>'


async def run_classification(input_path: str, output_dir: str, progress_callback=None):
    """Run the full classification pipeline, reporting progress via callback."""
    load_dotenv(ENV_PATH, override=True)

    from src.llm_client import LLMClient
    from src.excel_handler import ExcelHandler
    from src.part_classifier import PartClassifier
    from src.web_scraper import WebScraper
    from src.attribute_extractor import AttributeExtractor
    from main import process_part

    llm = LLMClient()
    handler = ExcelHandler(input_path, output_dir)
    classifier = PartClassifier(llm)
    attr_extractor = AttributeExtractor(llm)

    parts = handler.read_parts()
    total = len(parts)
    results = []

    if progress_callback:
        progress_callback("status", f"Processing {total} parts using {llm.display_name()}...")

    async with WebScraper() as scraper:
        for i, part in enumerate(parts):
            mfg_part_num = str(part.get("Manufacturer Part Number") or "").strip()
            part_name = str(part.get("Part Name") or "").strip()

            if progress_callback:
                progress_callback("progress", (i, total, mfg_part_num, part_name))

            result = await process_part(part, scraper, classifier, attr_extractor)
            results.append(result)

            if progress_callback:
                progress_callback("part_done", {
                    "index": i + 1,
                    "total": total,
                    "part_num": mfg_part_num,
                    "part_name": part_name,
                    "part_class": result["part_class"],
                    "source_url": result["source_url"] or "part name",
                    "attr_count": len(result["attributes"]),
                    "attributes": result["attributes"],
                })

    written = handler.write_class_files(results)
    return results, written


# ==========================================================================
# STREAMLIT APP
# ==========================================================================

st.set_page_config(
    page_title="Part Classification Agent",
    page_icon="https://raw.githubusercontent.com/streamlit/streamlit/develop/lib/streamlit/static/favicon.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

inject_custom_css(dark_mode=True)

# -- Sidebar ---------------------------------------------------------------
env = read_env()
current_provider = env.get("LLM_PROVIDER", "").lower()
current_key = env.get("LLM_API_KEY", env.get("GROQ_API_KEY", ""))
current_model = env.get("LLM_MODEL", "")

configured = bool(
    (env.get("LLM_PROVIDER") and env.get("LLM_API_KEY"))
    or env.get("GROQ_API_KEY")
    or env.get("LLM_PROVIDER", "").lower() == "ollama"
)

with st.sidebar:
    # Sidebar branding
    st.markdown("""
    <div class="sidebar-logo">
        <div class="sidebar-logo-icon">&#9881;</div>
        <div>
            <div class="sidebar-logo-text">Part Classifier</div>
            <div class="sidebar-logo-sub">AI Classification Agent</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-label">LLM Provider</div>', unsafe_allow_html=True)

    provider_labels = list(PROVIDER_DISPLAY.keys())
    default_idx = 0
    for i, label in enumerate(provider_labels):
        if PROVIDER_DISPLAY[label] == current_provider:
            default_idx = i
            break

    selected_label = st.selectbox(
        "Provider",
        provider_labels,
        index=default_idx,
        label_visibility="collapsed",
    )
    provider_key = PROVIDER_DISPLAY[selected_label]
    preset = PROVIDER_PRESETS[provider_key]

    st.markdown('<div class="section-label">Model</div>', unsafe_allow_html=True)

    model_options = list(preset["models"])
    if current_model and current_model not in model_options:
        model_options = [current_model] + model_options
    default_model_idx = model_options.index(current_model) if current_model in model_options else 0
    selected_model = st.selectbox(
        "Model",
        model_options,
        index=default_model_idx,
        label_visibility="collapsed",
    )

    if preset["needs_key"]:
        st.markdown('<div class="section-label">API Key</div>', unsafe_allow_html=True)
        api_key_input = st.text_input(
            "API Key",
            value=current_key if current_provider == provider_key else "",
            type="password",
            placeholder=preset["key_hint"],
            label_visibility="collapsed",
        )
        if preset["key_url"]:
            st.caption(f"[Get your API key here]({preset['key_url']})")
    else:
        api_key_input = ""
        st.info("Ollama runs locally. Make sure `ollama serve` is running.")

    if st.button("Save Configuration", type="primary", use_container_width=True):
        if preset["needs_key"] and not api_key_input.strip():
            st.error("API key is required.")
        else:
            write_env(provider_key, api_key_input.strip(), selected_model)
            st.success("Configuration saved!")
            st.rerun()

    # Status display
    st.markdown('<hr style="margin:8px 0;border-color:rgba(148,163,184,0.15);">', unsafe_allow_html=True)
    st.markdown('<div class="section-label">Status</div>', unsafe_allow_html=True)

    if configured:
        p_key = env.get("LLM_PROVIDER", "groq")
        p_model = env.get("LLM_MODEL", preset["default_model"])
        p_emoji = PROVIDER_EMOJI.get(p_key, "&#9881;")
        st.markdown(f"""
        <div class="provider-status">
            <div class="provider-status-row">
                <div class="provider-status-icon">{p_emoji}</div>
                <div class="provider-status-info">
                    <div class="provider-status-name">{PROVIDER_PRESETS.get(p_key, {}).get('label', p_key)}</div>
                    <div class="provider-status-model">{p_model}</div>
                </div>
                <span class="status-pill active"><span class="dot"></span> Ready</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="provider-status">
            <div class="provider-status-row">
                <span class="status-pill inactive"><span class="dot"></span> Not Configured</span>
            </div>
        </div>
        """, unsafe_allow_html=True)


# -- Main Area -------------------------------------------------------------

# Hero banner
st.markdown("""
<div class="hero-banner">
    <p class="hero-title">Part Classification Agent</p>
    <p class="hero-subtitle">Classify mechanical parts, scrape specifications from the web, and generate structured Excel output -- powered by AI.</p>
    <div class="hero-badge">&#9889; Supports Groq, OpenAI, Claude, Ollama</div>
</div>
""", unsafe_allow_html=True)

if not configured:
    st.markdown("""
    <div class="ui-card" style="text-align:center; padding:48px;">
        <div style="font-size:2.5rem; margin-bottom:12px;">&#9881;</div>
        <h3 style="color:#f1f5f9; margin-bottom:8px;">Get Started</h3>
        <p style="color:#94a3b8; max-width:400px; margin:0 auto;">
            Configure your LLM provider in the sidebar to begin classifying parts.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# -- Input Section ---------------------------------------------------------
col1, col2 = st.columns([2, 1])

with col1:
    input_folder = st.text_input(
        "Input Folder",
        value=DEFAULT_INPUT_DIR,
        help="Folder containing .xlsx input files",
    )

with col2:
    output_folder = st.text_input(
        "Output Folder",
        value=str(OUTPUT_DIR),
        help="Where output files will be saved",
    )

excel_files = get_excel_files(input_folder)

if not excel_files:
    st.warning(f"No .xlsx files found in `{input_folder}`")
    st.stop()

selected_file = st.selectbox(
    "Select Input File",
    excel_files,
    format_func=lambda p: p.name,
)

# Preview
if selected_file:
    import openpyxl
    wb = openpyxl.load_workbook(str(selected_file))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))
    wb.close()

    # Stats row
    unique_mfrs = len(set(r.get("Manufacturer Name", "") for r in rows))
    units = set(str(r.get("Unit of Measure", "")).lower() for r in rows)
    unit_str = " / ".join(sorted(u for u in units if u))

    st.markdown(f"""
    <div class="stat-row">
        <div class="stat-box">
            <div class="stat-value blue">{len(rows)}</div>
            <div class="stat-label">Total Parts</div>
        </div>
        <div class="stat-box">
            <div class="stat-value purple">{unique_mfrs}</div>
            <div class="stat-label">Manufacturers</div>
        </div>
        <div class="stat-box">
            <div class="stat-value green">{unit_str or 'N/A'}</div>
            <div class="stat-label">Unit Systems</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander(f"Preview input data ({len(rows)} rows)", expanded=False):
        st.dataframe(rows, use_container_width=True, hide_index=True)

# Run button
run_clicked = st.button(
    "Run Classification",
    type="primary",
    use_container_width=True,
    disabled=not selected_file,
)

# -- Classification Run ----------------------------------------------------

if run_clicked:
    st.markdown("---")

    progress_bar = st.progress(0)
    status_text = st.empty()
    part_results = []

    def progress_callback(event, data):
        if event == "status":
            status_text.info(data)
        elif event == "progress":
            i, total, part_num, part_name = data
            pct = int((i / total) * 100)
            progress_bar.progress(pct, text=f"Processing {i+1}/{total}: {part_num} - {part_name}")
        elif event == "part_done":
            part_results.append(data)
            pct = int((data["index"] / data["total"]) * 100)
            progress_bar.progress(pct, text=f"Completed {data['index']}/{data['total']}")

    try:
        results, written_files = asyncio.run(
            run_classification(str(selected_file), output_folder, progress_callback)
        )

        progress_bar.progress(100, text="Complete!")
        status_text.empty()

        # -- Success banner ------------------------------------------------
        total_attrs = sum(pr["attr_count"] for pr in part_results)
        unique_classes = len(set(pr["part_class"] for pr in part_results))

        st.markdown(f"""
        <div class="ui-card" style="background: linear-gradient(135deg, rgba(22,163,98,0.1), rgba(16,185,129,0.08)); border-color: rgba(34,197,94,0.3);">
            <div class="ui-card-header">
                <div class="ui-card-icon green" style="background:rgba(22,163,98,0.2);">&#10003;</div>
                <div>
                    <p class="ui-card-title" style="color:#16a34a;">Classification Complete</p>
                    <p class="ui-card-desc">All parts have been processed successfully</p>
                </div>
            </div>
            <div class="stat-row">
                <div class="stat-box" style="background:#0f172a;">
                    <div class="stat-value blue">{len(results)}</div>
                    <div class="stat-label">Parts Processed</div>
                </div>
                <div class="stat-box" style="background:#0f172a;">
                    <div class="stat-value purple">{unique_classes}</div>
                    <div class="stat-label">Classes Found</div>
                </div>
                <div class="stat-box" style="background:#0f172a;">
                    <div class="stat-value green">{total_attrs}</div>
                    <div class="stat-label">Attributes Extracted</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # -- Download section ----------------------------------------------
        st.markdown("""
        <div class="ui-card">
            <div class="ui-card-header">
                <div class="ui-card-icon purple">&#128229;</div>
                <div>
                    <p class="ui-card-title">Output Files</p>
                    <p class="ui-card-desc">Download the generated Excel files per class</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        cols = st.columns(min(len(written_files), 3))
        for idx, f in enumerate(written_files):
            fname = Path(f).name
            with cols[idx % len(cols)]:
                with open(f, "rb") as fh:
                    st.download_button(
                        label=f"&#128196; {fname}",
                        data=fh.read(),
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

        # -- Per-part results ----------------------------------------------
        st.markdown("""
        <div class="ui-card">
            <div class="ui-card-header">
                <div class="ui-card-icon amber">&#128269;</div>
                <div>
                    <p class="ui-card-title">Results Detail</p>
                    <p class="ui-card-desc">Click any part to see extracted attributes</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        for pr in part_results:
            with st.expander(
                f"{pr['part_num']}  |  {pr['part_class']}  |  {pr['attr_count']} attributes",
                expanded=False,
            ):
                c1, c2, c3 = st.columns(3)
                c1.metric("Part Class", pr["part_class"])
                c2.metric("Attributes", pr["attr_count"])
                c3.metric("Source", "Web" if pr["source_url"] != "part name" else "Part Name")

                if pr["source_url"] and pr["source_url"] != "part name":
                    st.caption(f"Source: {pr['source_url']}")

                if pr["attributes"]:
                    st.markdown(
                        render_attr_table(pr["attributes"]),
                        unsafe_allow_html=True,
                    )

    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        st.error(f"Error: {e}")
        import traceback
        with st.expander("Full error details"):
            st.code(traceback.format_exc())

# -- Footer ----------------------------------------------------------------
st.markdown("""
<div class="app-footer">
    <div class="footer-line"></div>
    <p>Developed by <strong>Anbarasu M</strong> &nbsp;&middot;&nbsp;
    <a href="https://digitalthread.ai" target="_blank">digitalthread.ai</a></p>
</div>
<style>
.app-footer {
    text-align: center;
    padding: 24px 0 12px;
    margin-top: 40px;
}
.app-footer .footer-line {
    width: 60px;
    height: 2px;
    background: linear-gradient(90deg, transparent, #6366f1, transparent);
    margin: 0 auto 14px;
    border-radius: 2px;
}
.app-footer p {
    font-size: 0.78rem;
    color: #64748b !important;
    letter-spacing: 0.02em;
}
.app-footer a {
    color: #818cf8 !important;
    text-decoration: none;
    font-weight: 600;
    transition: color 0.2s;
}
.app-footer a:hover {
    color: #a5b4fc !important;
}
</style>
""", unsafe_allow_html=True)
