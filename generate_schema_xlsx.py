"""Generate ClassificationSchema.xlsx from the current hardcoded data in attr_schema.py."""
import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()

# === Sheet 1: Classes ===
ws_classes = wb.active
ws_classes.title = "Classes"

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col, h in enumerate(["Class Name", "TC Class ID", "Description"], 1):
    cell = ws_classes.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

classes = [
    ("Washer", "", "General washer category"),
    ("Lock Washer", "", "General lock washer"),
    ("Split Lock Washer", "", "Helical spring lock washers (DIN 127)"),
    ("Flat Washer", "", "Standard flat washers"),
    ("Fender Washer", "", "Large OD flat washers"),
    ("Internal Tooth Lock Washer", "", "Internal tooth lock washers"),
    ("External Tooth Lock Washer", "", "External tooth lock washers"),
    ("Nut", "", "General nut category"),
    ("Lock Nut", "", "Prevailing torque lock nuts"),
    ("Hex Nut", "", "Standard hex nuts"),
    ("Wing Nut", "", "Wing/butterfly nuts"),
    ("Bolt", "", "General bolt category"),
    ("Hex Bolt", "", "Standard hex head bolts"),
    ("Carriage Bolt", "", "Round head square neck bolts"),
    ("Screw", "", "General screw category"),
    ("Cap Screw", "", "Hex socket cap screws"),
    ("Set Screw", "", "Headless set screws"),
    ("Machine Screw", "", "Machine screws"),
    ("Socket Head Cap Screw", "", "Socket head cap screws (SHCS)"),
    ("Hook", "", "Hooks"),
    ("Eye Bolt", "", "Eye bolts"),
    ("Eye Hook", "", "Eye hooks"),
    ("Pin", "", "General pin category"),
    ("Cotter Pin", "", "Split cotter pins"),
    ("Dowel Pin", "", "Cylindrical dowel pins"),
    ("Roll Pin", "", "Spring/roll pins"),
    ("Rivet", "", "General rivet category"),
    ("Blind Rivet", "", "Pop/blind rivets"),
    ("Clip", "", "General clip category"),
    ("E-Clip", "", "External retaining clips"),
    ("C-Clip", "", "C-shaped retaining clips"),
    ("Ring", "", "General ring category"),
    ("Retaining Ring", "", "Circlips/snap rings"),
    ("Bushing", "", "Bushings/sleeves"),
    ("Spacer", "", "Spacers"),
    ("Standoff", "", "Threaded standoffs"),
    ("Stud", "", "Threaded studs"),
    ("Insert", "", "Threaded inserts"),
    ("Anchor", "", "Anchors"),
    ("Spring", "", "General springs"),
    ("Compression Spring", "", "Compression springs"),
    ("Bracket", "", "Mounting brackets"),
    ("O-Ring", "", "Elastomer O-ring seals"),
    ("Seal", "", "General seals"),
    ("Gasket", "", "Gaskets"),
    ("Tube Fitting", "", "Compression tube fittings"),
    ("VCR Fitting", "", "Metal gasket face seal fittings"),
    ("Pipe Fitting", "", "Pipe fittings (NPT/BSP)"),
    ("Solenoid Valve", "", "Electrically actuated valves"),
    ("Pneumatic Valve", "", "Air-operated valves"),
    ("Pneumatic Cylinder", "", "Air cylinders/actuators"),
    ("Flow Controller", "", "Flow control devices"),
    ("Pressure Regulator", "", "Pressure regulators"),
    ("Ball Bearing", "", "General ball bearings"),
    ("Deep Groove Ball Bearing", "", "Single-row radial ball bearings"),
    ("Angular Contact Bearing", "", "Angular contact ball bearings"),
    ("Needle Bearing", "", "Needle roller bearings"),
    ("Crossed Roller Bearing", "", "Crossed roller bearings"),
    ("Linear Guide", "", "Linear motion guides/rails"),
    ("Linear Block", "", "Linear guide blocks/carriages"),
    ("Ball Screw", "", "Precision ball screws"),
    ("Proximity Sensor", "", "Inductive/capacitive proximity sensors"),
    ("Photoelectric Sensor", "", "Photoelectric sensors"),
    ("Fiber Optic Sensor", "", "Fiber optic sensors/amplifiers"),
    ("Laser Sensor", "", "Laser displacement/distance sensors"),
    ("Pressure Sensor", "", "Pressure transducers/gauges"),
    ("Connector", "", "Electrical connectors"),
    ("Terminal", "", "Terminal blocks/strips"),
    ("Relay", "", "Electromechanical relays"),
    ("Timer", "", "Industrial timers"),
    ("Vacuum Valve", "", "Vacuum isolation valves"),
    ("Gate Valve", "", "Gate/slide valves"),
    ("Vacuum Pump Accessory", "", "Vacuum pump accessories"),
    ("Wafer Shipper", "", "Wafer shipping containers"),
    ("Wafer Carrier", "", "Wafer transport carriers"),
    ("Filter", "", "General filters"),
    ("Gas Filter", "", "Gas-phase filters"),
    ("Liquid Filter", "", "Liquid-phase filters"),
    ("Mass Flow Controller", "", "MFC devices"),
    ("Pressure Gauge", "", "Pressure measurement gauges"),
    ("Vacuum Gauge", "", "Vacuum measurement gauges"),
]

for i, (name, tc_id, desc) in enumerate(classes, 2):
    ws_classes.cell(row=i, column=1, value=name)
    ws_classes.cell(row=i, column=2, value=tc_id)
    ws_classes.cell(row=i, column=3, value=desc)

ws_classes.column_dimensions["A"].width = 30
ws_classes.column_dimensions["B"].width = 15
ws_classes.column_dimensions["C"].width = 45

# === Sheet 2: Attributes ===
ws_attrs = wb.create_sheet("Attributes")

for col, h in enumerate(["Attribute Name", "Applicable Classes", "Aliases"], 1):
    cell = ws_attrs.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

# Class groups for readability
W = "Washer; Lock Washer; Split Lock Washer; Flat Washer; Fender Washer; Internal Tooth Lock Washer; External Tooth Lock Washer"
N = "Nut; Lock Nut; Hex Nut; Wing Nut"
BS = "Bolt; Hex Bolt; Carriage Bolt; Screw; Cap Screw; Set Screw; Machine Screw; Socket Head Cap Screw"
P = "Pin; Cotter Pin; Dowel Pin; Roll Pin"
BR = "Ball Bearing; Deep Groove Ball Bearing; Angular Contact Bearing; Needle Bearing; Crossed Roller Bearing"
LN = "Linear Guide; Linear Block; Ball Screw"
SL = "O-Ring; Seal; Gasket"
FT = "Tube Fitting; VCR Fitting; Pipe Fitting"
PN = "Solenoid Valve; Pneumatic Valve; Pneumatic Cylinder; Flow Controller; Pressure Regulator"
SN = "Proximity Sensor; Photoelectric Sensor; Fiber Optic Sensor; Laser Sensor; Pressure Sensor"
EL = "Connector; Terminal; Relay; Timer"
VC = "Vacuum Valve; Gate Valve; Vacuum Pump Accessory; Wafer Shipper; Wafer Carrier; Filter; Gas Filter; Liquid Filter; Mass Flow Controller; Pressure Gauge; Vacuum Gauge"

attributes = [
    # Universal
    ("Material", "*", "material type; alloy; body material; material - body"),
    ("Finish", "*", "coating; surface finish; plating; surface treatment"),
    ("Standard", "*", "specifications met; specs met; specification; spec; norm"),
    ("System of Measurement", "*", "measurement system; unit system"),
    ("RoHS", "*", "rohs compliant; rohs compliance"),
    ("REACH", "*", "reach compliant; reach compliance"),
    ("Performance", "*", "corrosion resistance"),
    ("Weight", "*", "mass; net weight"),
    # Dimensions
    ("Inner Diameter", f"{W}; {SL}; {BR}; {PN}; {VC}", "id; i.d.; bore diameter; bore; hole diameter; hole size; inside diameter"),
    ("Outer Diameter", f"{W}; {SL}; {BR}; {FT}", "od; o.d.; outside diameter; outer dia"),
    ("Thickness", f"{W}; {SL}; {BR}", "thk; thickness range; overall thickness"),
    ("Length", f"{BS}; {P}; {PN}; {LN}; {SN}; {FT}", "overall length; stroke; stroke length"),
    ("Width", f"{LN}; {SN}; {EL}", ""),
    ("Height", f"{N}; {LN}; {SN}; {EL}", ""),
    ("Diameter", f"{P}", "pin diameter; shaft diameter"),
    # Thread
    ("Screw Size", f"{W}; {N}; {BS}", "for screw size; thread size; thread; nominal thread size; bolt size; fastener size; screw/bolt size"),
    ("Thread Pitch", f"{N}; {BS}", "thread pitch - metric; threads per inch; pitch"),
    # Fastener
    ("Hardness", f"{W}; {N}; {BS}; {P}; {BR}; {SL}", "hardness rating"),
    ("Grade", f"{BS}", "class; strength grade; property class"),
    ("Drive Type", f"{BS}", "drive style; socket type"),
    ("Head Diameter", f"{BS}", "head dia"),
    ("Head Height", f"{BS}; {N}", "head thickness"),
    ("Width Across Flats", f"{N}; Hex Bolt", "wrench size; af; s dimension"),
    ("Point Type", "Set Screw", "point style; tip type"),
    ("Washer Type", f"{W}", "type; head type; head style"),
    ("Nut Type", f"{N}", "nut style"),
    ("Tolerance", f"{P}; {BR}; {LN}", "tolerance class; accuracy"),
    # Electrical
    ("Voltage", f"{PN}; {SN}; {EL}", "rated voltage; supply voltage; operating voltage"),
    ("Current", f"{SN}; {EL}", "rated current; max current; current rating"),
    ("Power", f"{PN}; {SN}; {EL}", "power consumption; wattage"),
    ("Resistance", f"{EL}", "contact resistance; insulation resistance"),
    # Pneumatic
    ("Pressure", f"{PN}; {FT}; {VC}", "operating pressure; rated pressure; working pressure; max pressure"),
    ("Temperature", f"{PN}; {SL}; {SN}; {VC}", "operating temperature; temperature range; max temperature"),
    ("Flow Rate", f"{PN}; {FT}; {VC}", "flow; cv value; conductance"),
    ("Port Size", f"{PN}; {FT}", "connection size; pipe size"),
    # Sensor
    ("Sensing Distance", f"{SN}", "detection range; sensing range; detecting distance"),
    ("Response Time", f"{SN}; {VC}", "switching frequency; response frequency"),
    ("Output Type", f"{SN}", "output; output configuration; npn; pnp"),
    ("Protection Rating", f"{SN}; {PN}", "ip rating; ingress protection; ip67; ip68"),
    # Bearing
    ("Dynamic Load Rating", f"{BR}; {LN}", "dynamic load; basic dynamic load"),
    ("Static Load Rating", f"{BR}; {LN}", "static load; basic static load"),
    ("Speed Rating", f"{BR}", "max speed; limiting speed; reference speed"),
    # Connector
    ("Number of Contacts", f"{EL}", "number of positions; number of pins; pin count; contacts"),
    ("Pitch", f"{EL}", "pin pitch; contact pitch; spacing"),
    ("Contact Material", f"{EL}", "contact plating; contact finish"),
    # Vacuum
    ("Micron Rating", "Filter; Gas Filter; Liquid Filter", "filtration rating; pore size; removal rating"),
]

for i, (name, classes_str, aliases) in enumerate(attributes, 2):
    ws_attrs.cell(row=i, column=1, value=name)
    ws_attrs.cell(row=i, column=2, value=classes_str)
    ws_attrs.cell(row=i, column=3, value=aliases)

ws_attrs.column_dimensions["A"].width = 25
ws_attrs.column_dimensions["B"].width = 80
ws_attrs.column_dimensions["C"].width = 80

wb.save("input/ClassificationSchema.xlsx")
print(f"Created: {len(classes)} classes, {len(attributes)} attributes")
