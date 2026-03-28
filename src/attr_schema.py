"""
Canonical attribute names and normalization for each part class.

SCHEMA SOURCE PRIORITY
----------------------
1. JSON files: input/Classes.json + input/Attributes.json  (preferred)
2. Excel file: input/ClassificationSchema.xlsx             (fallback)
3. Hardcoded defaults                                      (last resort)

PUBLIC API
----------
- KNOWN_CLASSES: list[str]           — all valid part class names
- TC_CLASS_IDS: dict[str, str]       — class name -> Teamcenter class ID
- CLASS_SCHEMAS: dict[str, list[str]]— class -> ordered canonical attribute names (inherited)
- ALIASES: dict[str, str]            — lowercase alias -> canonical attribute name
- CLASS_ALIASES: dict[str, str]      — lowercase class alias -> canonical class name
- CLASS_TREE_CHILDREN: dict[str, list[str]] — parent class name -> child class names
- LOV_MAP: dict[str, list[str]]      — canonical attr name -> LOV values
- ATTR_DICT: dict[str, dict]         — attr id -> full attribute record
- get_schema(part_class) -> list[str]
- get_tc_class_id(part_class) -> str
- normalize_attrs(raw, part_class) -> dict
- schema_source() -> str
"""

import json
import re
import warnings
from pathlib import Path
from typing import Any

import openpyxl


# ── File locations ───────────────────────────────────────────────────────────

_INPUT_DIR = Path(__file__).parent.parent / "input"
_CLASSES_JSON = _INPUT_DIR / "Classes.json"
_ATTRS_JSON = _INPUT_DIR / "Attributes.json"
_SCHEMA_XLSX = _INPUT_DIR / "ClassificationSchema.xlsx"


# ── Module-level data (populated at import time) ────────────────────────────

KNOWN_CLASSES: list[str] = []
TC_CLASS_IDS: dict[str, str] = {}          # class_name -> Teamcenter class ID
CLASS_SCHEMAS: dict[str, list[str]] = {}   # class_name -> [canonical attr names] (inherited)
ALIASES: dict[str, str] = {}               # lowercase alias -> canonical attr name
CLASS_ALIASES: dict[str, str] = {}         # lowercase class alias -> canonical class name
CLASS_TREE_CHILDREN: dict[str, list[str]] = {}  # parent name -> [child names]
LOV_MAP: dict[str, list[str]] = {}         # canonical attr name -> LOV values
ATTR_DICT: dict[str, dict] = {}            # attr id -> full attribute record
_DEFAULT_SCHEMA: list[str] = []
_SCHEMA_SOURCE: str = "none"               # "json", "excel", or "hardcoded"


# ── JSON loader ──────────────────────────────────────────────────────────────

def _load_from_json(classes_path: Path, attrs_path: Path) -> None:
    """Load schema from Classes.json + Attributes.json."""
    global KNOWN_CLASSES, TC_CLASS_IDS, CLASS_SCHEMAS, ALIASES, CLASS_ALIASES
    global CLASS_TREE_CHILDREN, LOV_MAP, ATTR_DICT, _DEFAULT_SCHEMA, _SCHEMA_SOURCE

    # --- Load attributes ---
    with open(attrs_path, "r", encoding="utf-8") as f:
        attrs_data = json.load(f)

    attr_by_id: dict[str, dict] = {}
    attr_name_by_id: dict[str, str] = {}

    for attr in attrs_data["attributes"]:
        aid = str(attr["id"])
        attr_by_id[aid] = attr
        attr_name_by_id[aid] = attr["name"]

        # Populate ATTR_DICT
        ATTR_DICT[aid] = attr

        # Populate ALIASES: canonical name + all aliases
        canonical = attr["name"]
        ALIASES[canonical.lower()] = canonical
        for alias in attr.get("aliases", []):
            ALIASES[alias.lower()] = canonical

        # Populate LOV_MAP
        if attr.get("lov"):
            LOV_MAP[canonical] = attr["lov"]

    # --- Load classes ---
    with open(classes_path, "r", encoding="utf-8") as f:
        classes_data = json.load(f)

    # Flatten the class tree, computing inherited attributes
    _flatten_tree(classes_data["classes"], inherited_attrs=[], attr_name_by_id=attr_name_by_id)

    # Build default schema from common attributes
    _DEFAULT_SCHEMA = [
        "Screw Size", "Inner Diameter", "Outer Diameter", "Thickness", "Length",
        "Width", "Height", "Material", "Finish", "Hardness", "Standard",
        "System of Measurement", "Performance", "RoHS", "REACH",
    ]

    _SCHEMA_SOURCE = "json"


def _flatten_tree(
    nodes: list[dict],
    inherited_attrs: list[str],
    attr_name_by_id: dict[str, str],
) -> None:
    """Recursively walk the class tree, populating module-level data.

    Each node accumulates inherited attributes from ancestors plus its own
    directly-assigned attributes. Only leaf and named intermediate classes
    are added to KNOWN_CLASSES.
    """
    for node in nodes:
        name = node["name"]
        class_id = node.get("id", "")
        children = node.get("children", [])
        aliases = node.get("aliases", [])

        # Resolve this node's directly-assigned attribute IDs to names
        direct_attr_names = []
        for aid in node.get("attributeslist", []):
            aname = attr_name_by_id.get(str(aid))
            if aname:
                direct_attr_names.append(aname)

        # Full attribute list = inherited + direct (no duplicates, order preserved)
        full_attrs = list(inherited_attrs)
        for a in direct_attr_names:
            if a not in full_attrs:
                full_attrs.append(a)

        # Register this class
        KNOWN_CLASSES.append(name)
        TC_CLASS_IDS[name] = class_id

        # Schema = full inherited attribute list
        CLASS_SCHEMAS[name] = full_attrs

        # Class aliases
        for alias in aliases:
            CLASS_ALIASES[alias.lower()] = name

        # Parent-child relationships
        if children:
            CLASS_TREE_CHILDREN[name] = [c["name"] for c in children]

        # Recurse into children, passing accumulated attrs
        if children:
            _flatten_tree(children, full_attrs, attr_name_by_id)


# ── Excel loader (fallback) ─────────────────────────────────────────────────

def _load_from_excel(path: Path) -> None:
    """Load schema from ClassificationSchema.xlsx."""
    global KNOWN_CLASSES, TC_CLASS_IDS, CLASS_SCHEMAS, ALIASES, _DEFAULT_SCHEMA, _SCHEMA_SOURCE

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # --- Sheet 1: Classes ---
    if "Classes" not in wb.sheetnames:
        raise ValueError("Sheet 'Classes' not found in ClassificationSchema.xlsx")

    ws_classes = wb["Classes"]
    for row in ws_classes.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        tc_id = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if name:
            KNOWN_CLASSES.append(name)
            if tc_id:
                TC_CLASS_IDS[name] = tc_id

    # --- Sheet 2: Attributes ---
    if "Attributes" not in wb.sheetnames:
        raise ValueError("Sheet 'Attributes' not found in ClassificationSchema.xlsx")

    ws_attrs = wb["Attributes"]
    for row in ws_attrs.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        attr_name = str(row[0]).strip()
        classes_str = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        aliases_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if not attr_name:
            continue

        # Parse applicable classes
        if classes_str == "*":
            applicable = set(KNOWN_CLASSES)
        else:
            applicable = {c.strip() for c in classes_str.split(";") if c.strip()}
            for cls in applicable:
                if cls not in KNOWN_CLASSES:
                    warnings.warn(
                        f"ClassificationSchema.xlsx: Attribute '{attr_name}' references "
                        f"unknown class '{cls}' (not in Classes sheet)"
                    )

        for cls in applicable:
            if cls in KNOWN_CLASSES:
                CLASS_SCHEMAS.setdefault(cls, []).append(attr_name)

        ALIASES[attr_name.lower()] = attr_name
        if aliases_str:
            for alias in aliases_str.split(";"):
                alias = alias.strip().lower()
                if alias:
                    ALIASES[alias] = attr_name

    wb.close()

    _DEFAULT_SCHEMA = [
        "Screw Size", "Inner Diameter", "Outer Diameter", "Thickness", "Length",
        "Width", "Height", "Material", "Finish", "Hardness", "Standard",
        "System of Measurement", "Performance", "RoHS", "REACH",
    ]

    _SCHEMA_SOURCE = "excel"


# ── Hardcoded fallback ───────────────────────────────────────────────────────

def _load_hardcoded_defaults() -> None:
    """Fallback: populate from hardcoded data when JSON and Excel are missing."""
    global KNOWN_CLASSES, TC_CLASS_IDS, CLASS_SCHEMAS, ALIASES, CLASS_ALIASES
    global CLASS_TREE_CHILDREN, _DEFAULT_SCHEMA, _SCHEMA_SOURCE

    KNOWN_CLASSES.extend([
        # Fasteners
        "Washer", "Lock Washer", "Split Lock Washer", "Flat Washer",
        "Fender Washer", "Internal Tooth Lock Washer", "External Tooth Lock Washer",
        "Nut", "Lock Nut", "Hex Nut", "Wing Nut",
        "Bolt", "Hex Bolt", "Carriage Bolt",
        "Screw", "Cap Screw", "Set Screw", "Machine Screw", "Socket Head Cap Screw",
        "Hook", "Eye Bolt", "Eye Hook",
        "Pin", "Cotter Pin", "Dowel Pin", "Roll Pin",
        "Rivet", "Blind Rivet",
        "Clip", "E-Clip", "C-Clip",
        "Ring", "Retaining Ring",
        "Bushing", "Spacer", "Standoff",
        "Stud", "Insert", "Anchor",
        "Spring", "Compression Spring", "Bracket",
        # Seals & Fittings
        "O-Ring", "Seal", "Gasket",
        "Tube Fitting", "VCR Fitting", "Pipe Fitting",
        # Pneumatics & Hydraulics
        "Solenoid Valve", "Pneumatic Valve", "Pneumatic Cylinder",
        "Flow Controller", "Pressure Regulator",
        # Bearings & Linear Motion
        "Ball Bearing", "Deep Groove Ball Bearing", "Angular Contact Bearing",
        "Needle Bearing", "Crossed Roller Bearing",
        "Linear Guide", "Linear Block", "Ball Screw",
        # Sensors & Electrical
        "Proximity Sensor", "Photoelectric Sensor", "Fiber Optic Sensor",
        "Laser Sensor", "Pressure Sensor",
        "Connector", "Terminal", "Relay", "Timer",
        # Vacuum & Semiconductor
        "Vacuum Valve", "Gate Valve", "Vacuum Pump Accessory",
        "Wafer Shipper", "Wafer Carrier",
        "Filter", "Gas Filter", "Liquid Filter",
        "Mass Flow Controller", "Pressure Gauge", "Vacuum Gauge",
    ])

    CLASS_SCHEMAS.update({
        "Flat Washer": [
            "Screw Size", "Inner Diameter", "Outer Diameter", "Thickness",
            "Material", "Finish", "Hardness", "Standard",
            "System of Measurement", "Washer Type", "Performance",
        ],
        "Split Lock Washer": [
            "Screw Size", "Inner Diameter", "Outer Diameter", "Thickness",
            "Material", "Finish", "Hardness", "Standard",
            "System of Measurement", "Washer Type", "Performance",
            "RoHS", "REACH",
        ],
        "Hex Nut": [
            "Thread Size", "Thread Pitch", "Width Across Flats", "Height",
            "Material", "Finish", "Hardness", "Standard",
            "System of Measurement", "Nut Type",
        ],
        "Hex Bolt": [
            "Thread Size", "Thread Pitch", "Length", "Width Across Flats",
            "Head Height", "Material", "Finish", "Hardness", "Grade", "Standard",
            "System of Measurement",
        ],
        "Cap Screw": [
            "Thread Size", "Thread Pitch", "Length", "Head Diameter",
            "Head Height", "Drive Type", "Material", "Finish", "Hardness",
            "Grade", "Standard", "System of Measurement",
        ],
    })

    ALIASES.update({
        "screw size": "Screw Size", "for screw size": "Screw Size",
        "thread size": "Screw Size", "thread": "Screw Size",
        "nominal thread size": "Screw Size", "bolt size": "Screw Size",
        "id": "Inner Diameter", "inner diameter": "Inner Diameter",
        "inner dia": "Inner Diameter", "bore diameter": "Inner Diameter",
        "bore": "Inner Diameter", "hole diameter": "Inner Diameter",
        "od": "Outer Diameter", "outer diameter": "Outer Diameter",
        "outside diameter": "Outer Diameter",
        "thk": "Thickness", "thickness": "Thickness",
        "height": "Height", "thickness range": "Thickness",
        "material": "Material", "material type": "Material",
        "finish": "Finish", "coating": "Finish", "surface finish": "Finish",
        "plating": "Finish", "surface treatment": "Finish",
        "standard": "Standard", "specifications met": "Standard",
        "spec": "Standard", "norm": "Standard",
        "system of measurement": "System of Measurement",
        "hardness": "Hardness", "hardness rating": "Hardness",
        "performance": "Performance",
        "washer type": "Washer Type", "type": "Washer Type",
        "rohs": "RoHS", "reach": "REACH",
        "screw/bolt size": "Screw Size", "inside diameter": "Inner Diameter",
        "overall thickness": "Thickness", "body material": "Material",
        "material - body": "Material", "drive style": "Drive Type",
        "drive type": "Drive Type", "overall length": "Length",
        "thread pitch - metric": "Thread Pitch",
        "threads per inch": "Thread Pitch",
    })

    CLASS_ALIASES.update({
        "dgbb": "Deep Groove Ball Bearing",
        "deep groove": "Deep Groove Ball Bearing",
        "radial ball bearing": "Deep Groove Ball Bearing",
        "angular contact": "Angular Contact Bearing",
        "needle roller": "Needle Bearing",
        "crossed roller": "Crossed Roller Bearing",
        "shcs": "Socket Head Cap Screw",
        "socket head cap": "Socket Head Cap Screw",
        "socket cap screw": "Socket Head Cap Screw",
        "split lock": "Split Lock Washer",
        "spring lock washer": "Split Lock Washer",
        "int tooth": "Internal Tooth Lock Washer",
        "ext tooth": "External Tooth Lock Washer",
        "fender": "Fender Washer",
        "hex cap screw": "Hex Bolt",
        "hex head bolt": "Hex Bolt",
        "carriage": "Carriage Bolt",
        "nylon insert": "Lock Nut",
        "nyloc": "Lock Nut",
        "pop rivet": "Blind Rivet",
        "cotter": "Cotter Pin",
        "dowel": "Dowel Pin",
        "roll pin": "Roll Pin",
        "spring pin": "Roll Pin",
        "e-clip": "E-Clip",
        "c-clip": "C-Clip",
        "snap ring": "Retaining Ring",
        "circlip": "Retaining Ring",
        "oring": "O-Ring",
        "o ring": "O-Ring",
        "compression fitting": "Tube Fitting",
        "vcr fitting": "VCR Fitting",
        "solenoid valve": "Solenoid Valve",
        "air cylinder": "Pneumatic Cylinder",
        "round cylinder": "Pneumatic Cylinder",
        "compact cylinder": "Pneumatic Cylinder",
        "inductive sensor": "Proximity Sensor",
        "fiber sensor": "Fiber Optic Sensor",
        "fiber unit": "Fiber Optic Sensor",
        "pressure transducer": "Pressure Sensor",
        "guide block": "Linear Block",
        "ball screw": "Ball Screw",
    })

    CLASS_TREE_CHILDREN.update({
        "Washer": ["Flat Washer", "Fender Washer", "Lock Washer"],
        "Lock Washer": ["Split Lock Washer", "Internal Tooth Lock Washer",
                        "External Tooth Lock Washer"],
        "Nut": ["Hex Nut", "Lock Nut", "Wing Nut"],
        "Bolt": ["Hex Bolt", "Carriage Bolt", "Eye Bolt"],
        "Screw": ["Cap Screw", "Set Screw", "Machine Screw"],
        "Cap Screw": ["Socket Head Cap Screw"],
        "Pin": ["Cotter Pin", "Dowel Pin", "Roll Pin"],
        "Rivet": ["Blind Rivet"],
        "Clip": ["E-Clip", "C-Clip"],
        "Ring": ["Retaining Ring"],
        "Hook": ["Eye Hook"],
        "Spring": ["Compression Spring"],
        "Seal": ["O-Ring", "Gasket"],
        "Fitting": ["Tube Fitting", "VCR Fitting", "Pipe Fitting"],
        "Bearing": ["Ball Bearing", "Needle Bearing", "Crossed Roller Bearing"],
        "Ball Bearing": ["Deep Groove Ball Bearing", "Angular Contact Bearing"],
        "Linear Motion": ["Linear Guide", "Linear Block", "Ball Screw"],
        "Sensor": ["Proximity Sensor", "Photoelectric Sensor", "Fiber Optic Sensor",
                    "Laser Sensor", "Pressure Sensor"],
        "Connector": ["Terminal", "Relay"],
        "Valve": ["Solenoid Valve", "Pneumatic Valve", "Vacuum Valve", "Gate Valve"],
        "Pneumatic": ["Pneumatic Cylinder", "Pressure Regulator", "Flow Controller"],
        "Filter": ["Gas Filter", "Liquid Filter"],
        "Gauge": ["Pressure Gauge", "Vacuum Gauge", "Mass Flow Controller"],
        "Wafer Handling": ["Wafer Carrier", "Wafer Shipper"],
    })

    _DEFAULT_SCHEMA = [
        "Screw Size", "Inner Diameter", "Outer Diameter", "Thickness", "Length",
        "Width", "Height", "Material", "Finish", "Hardness", "Standard",
        "System of Measurement", "Performance", "RoHS", "REACH",
    ]

    _SCHEMA_SOURCE = "hardcoded"


# ── Load schema on import ────────────────────────────────────────────────────

def _load_schema() -> None:
    """Load schema: JSON first, then Excel, then hardcoded defaults."""
    # Priority 1: JSON files
    if _CLASSES_JSON.exists() and _ATTRS_JSON.exists():
        try:
            _load_from_json(_CLASSES_JSON, _ATTRS_JSON)
            return
        except Exception as e:
            warnings.warn(f"Failed to load JSON schema ({e}), trying Excel fallback")

    # Priority 2: Excel
    if _SCHEMA_XLSX.exists():
        try:
            _load_from_excel(_SCHEMA_XLSX)
            return
        except Exception as e:
            warnings.warn(
                f"Failed to load ClassificationSchema.xlsx ({e}), using hardcoded defaults"
            )
    else:
        if not (_CLASSES_JSON.exists() and _ATTRS_JSON.exists()):
            warnings.warn(
                f"No schema files found (JSON or Excel), using hardcoded defaults"
            )

    # Priority 3: Hardcoded
    _load_hardcoded_defaults()


_load_schema()


# ── LOV normalization helper ─────────────────────────────────────────────────

def _normalize_key(s: str) -> str:
    """Lowercase, strip spaces, hyphens, underscores for fuzzy matching."""
    return re.sub(r"[\s\-_]", "", s.lower())


def _normalize_to_lov(value: str, lov_values: list[str]) -> str:
    """Match a raw value to a LOV entry using fuzzy comparison.

    "Stainless Steel" -> "StainlessSteel"
    "zinc plated" -> "ZincPlated"
    Returns the original value if no match found.
    """
    norm_val = _normalize_key(value)
    for lov_entry in lov_values:
        if _normalize_key(lov_entry) == norm_val:
            return lov_entry
    return value


# ── Public API ───────────────────────────────────────────────────────────────

def get_schema(part_class: str) -> list[str]:
    """Return canonical attribute list for the given class."""
    if part_class in CLASS_SCHEMAS:
        return CLASS_SCHEMAS[part_class]
    # Fuzzy match
    lower = part_class.lower()
    for key, schema in CLASS_SCHEMAS.items():
        if key.lower() in lower or lower in key.lower():
            return schema
    return _DEFAULT_SCHEMA


def get_tc_class_id(part_class: str) -> str:
    """Return Teamcenter class ID for the given class, or empty string."""
    if part_class in TC_CLASS_IDS:
        return TC_CLASS_IDS[part_class]
    lower = part_class.lower()
    for key, tc_id in TC_CLASS_IDS.items():
        if key.lower() == lower:
            return tc_id
    return ""


def _normalize_class_name(name: str) -> str:
    """Normalize a class name for fuzzy matching: lowercase, strip plural suffix."""
    n = name.lower().strip()
    # Strip common plural suffixes
    if n.endswith("ies"):
        n = n[:-3] + "y"
    elif n.endswith("es") and not n.endswith("ses"):
        n = n[:-2]
    elif n.endswith("s") and not n.endswith("ss"):
        n = n[:-1]
    return n


def map_to_json_class(detected_class: str) -> tuple[str, bool]:
    """Map a detected class name to the nearest Classes.json class.

    Returns (mapped_class, found_in_json).
    Matching is case-insensitive and grammar-insensitive (singular/plural).

    Example: "Flat Washer" not in JSON, but "WASHERS" is → returns ("WASHERS", True)
    """
    if not detected_class or detected_class in ("Unclassified", "Unknown", "Error"):
        return detected_class, False

    # Build normalized lookup from JSON classes
    json_classes = list(CLASS_SCHEMAS.keys())
    norm_map = {}  # normalized_name → original JSON class name
    for jc in json_classes:
        norm_map[_normalize_class_name(jc)] = jc

    # 1. Exact match (case-insensitive + plural-insensitive)
    detected_norm = _normalize_class_name(detected_class)
    if detected_norm in norm_map:
        return norm_map[detected_norm], True

    # 2. Check if any word in detected class matches a JSON class
    # e.g., "Flat Washer" → check "flat", "washer" against JSON classes
    detected_words = detected_class.lower().split()
    for word in detected_words:
        word_norm = _normalize_class_name(word)
        if word_norm in norm_map:
            return norm_map[word_norm], True

    # 3. Check if detected class is a substring of any JSON class or vice versa
    for jc_norm, jc_original in norm_map.items():
        if detected_norm in jc_norm or jc_norm in detected_norm:
            return jc_original, True

    # 4. Walk parent hierarchy — check if any parent class exists in JSON
    # (uses _PARENT_CHILDREN from class_extractor if available)
    try:
        from src.class_extractor import _CHILD_PARENTS
        current = detected_class
        visited = set()
        while current and current not in visited:
            visited.add(current)
            parents = _CHILD_PARENTS.get(current, set())
            for parent in parents:
                parent_norm = _normalize_class_name(parent)
                if parent_norm in norm_map:
                    return norm_map[parent_norm], True
            # Try next level up
            current = next(iter(parents), None) if parents else None
    except ImportError:
        pass

    # 5. No match — return as-is, not in JSON
    return detected_class, False


def schema_source() -> str:
    """Return 'json', 'excel', or 'hardcoded' indicating where schema was loaded from."""
    return _SCHEMA_SOURCE


def normalize_attrs(raw: dict, part_class: str) -> dict[str, str]:
    """
    1. Alias-normalize all keys in `raw`.
    2. LOV-normalize values where applicable.
    3. Order them according to the class schema (unknown keys go at the end).
    Returns a plain dict with consistent, canonical keys.
    """
    schema = get_schema(part_class)

    # Step 1: normalize keys and LOV values
    normalized: dict[str, str] = {}
    for k, v in raw.items():
        canonical = ALIASES.get(k.strip().lower(), k.strip())
        # Prefer schema-cased version if available
        for schema_key in schema:
            if schema_key.lower() == canonical.lower():
                canonical = schema_key
                break

        str_val = str(v).strip()

        # Step 2: LOV normalization
        if canonical in LOV_MAP and str_val:
            str_val = _normalize_to_lov(str_val, LOV_MAP[canonical])

        normalized[canonical] = str_val

    # Step 3: order by schema, then append any extras not in schema
    ordered: dict[str, str] = {}
    for key in schema:
        if key in normalized:
            ordered[key] = normalized[key]
    for key, val in normalized.items():
        if key not in ordered:
            ordered[key] = val

    return ordered
