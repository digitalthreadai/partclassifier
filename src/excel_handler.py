"""
Read parts from the input Excel file.
Write one output Excel file per part class into the output/ folder.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import Any


def _clean_header(h) -> str:
    """Remove invisible chars, BOM, non-breaking spaces from Excel headers.

    PLM/ERP exports often embed invisible characters in cell values.
    """
    if not h:
        return ""
    s = str(h)
    # Remove BOM, zero-width chars, non-breaking spaces
    s = s.replace('\ufeff', '').replace('\u200b', '').replace('\u200c', '')
    s = s.replace('\u200d', '').replace('\u00a0', ' ').replace('\ufffe', '')
    # Remove other control characters (except newline/tab)
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s)
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s).strip()
    return s

INPUT_COLUMNS = [
    "Part Number",
    "Part Name",
    "Manufacturer Part Number",
    "Manufacturer Name",
    "Unit of Measure",
]

# Fuzzy aliases for input column headers — handles different naming conventions
COLUMN_ALIASES = {
    "Part Number": ["PN", "Internal PN", "Item Number", "Item No", "Part No", "Part #"],
    "Part Name": ["Description", "Part Description", "Part Desc", "Name", "Item Name",
                   "Item Description", "Part Title", "Component Name", "Component Description"],
    "Manufacturer Part Number": ["Mfg Part Number", "MFG PN", "Mfg PN", "MPN",
                                  "Manufacturer PN", "Mfg Part No", "Mfr Part Number",
                                  "Mfr PN", "Vendor Part Number", "Vendor PN"],
    "Manufacturer Name": ["Mfg Name", "MFG Name", "Manufacturer", "Mfg", "Mfr Name",
                           "Mfr", "Vendor Name", "Vendor", "Supplier", "Supplier Name"],
    "Unit of Measure": ["UOM", "Unit", "Units", "Measure"],
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")       # Navy — input + prefix + TC attrs
AGENT_HEADER_FILL = PatternFill("solid", fgColor="4A4A6A") # Gray-purple — agent-extracted attrs
METRICS_HEADER_FILL = PatternFill("solid", fgColor="2D5F2D") # Dark green — quality metrics
LOV_MISMATCH_HEADER_FILL = PatternFill("solid", fgColor="C04040")  # Red — LOV mismatch flag
RANGE_ORIG_HEADER_FILL = PatternFill("solid", fgColor="2E6B8A")   # Steel blue — original range values
FRACTION_ORIG_HEADER_FILL = PatternFill("solid", fgColor="B85C00") # Amber — original fraction values
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", wrap_text=True)

# Conditional fill colors for percentage metrics
GREEN_FILL = PatternFill("solid", fgColor="1B5E20")    # ≥80%
AMBER_FILL = PatternFill("solid", fgColor="E65100")    # 50-79%
RED_FILL = PatternFill("solid", fgColor="B71C1C")      # <50%
LOV_MISMATCH_CELL_FILL = PatternFill("solid", fgColor="6B2020")  # Dark red — LOV mismatch cell

# Result columns written before the dynamic attribute columns
RESULT_PREFIX_COLS = ["Part Class", "TC Class ID", "Source URL", "Unit of Measure (Agent)"]

# Quality metric columns (after prefix, before TC attrs)
METRICS_COLS = [
    "Extraction Coverage %",
    "Source Reliability %",
    "Classification Confidence %",
    "Source Type",
    "LOV Compliance %",
    "Validation Action",
]
_PCT_METRICS = {"Extraction Coverage %", "Source Reliability %",
                "Classification Confidence %", "LOV Compliance %"}


class ExcelHandler:
    def __init__(self, input_path: str, output_dir: str):
        self.input_path = input_path
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def read_parts(self) -> list[dict[str, Any]]:
        """Return list of dicts keyed by INPUT_COLUMNS header names.

        Uses fuzzy column matching: tries exact match, then aliases,
        then case-insensitive match. Prints matched columns for debugging.
        """
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb.active
        raw_headers = [cell.value for cell in ws[1]]
        # Clean headers: remove invisible chars, BOM, collapse spaces
        headers = [_clean_header(h) for h in raw_headers]
        headers_lower = [h.lower() for h in headers]

        input_indices: dict[str, int] = {}
        for col in INPUT_COLUMNS:
            # Try exact match first
            if col in headers:
                input_indices[col] = headers.index(col)
                continue
            # Try aliases
            matched = False
            for alias in COLUMN_ALIASES.get(col, []):
                if alias in headers:
                    input_indices[col] = headers.index(alias)
                    print(f"  Column matched: '{alias}' -> '{col}'")
                    matched = True
                    break
            if matched:
                continue
            # Try case-insensitive match
            col_lower = col.lower()
            for i, h in enumerate(headers_lower):
                if h == col_lower:
                    input_indices[col] = i
                    print(f"  Column matched (case-insensitive): '{headers[i]}' -> '{col}'")
                    break
            # Try case-insensitive alias match
            if col not in input_indices:
                for alias in COLUMN_ALIASES.get(col, []):
                    alias_lower = alias.lower()
                    for i, h in enumerate(headers_lower):
                        if h == alias_lower:
                            input_indices[col] = i
                            print(f"  Column matched (alias): '{headers[i]}' -> '{col}'")
                            break
                    if col in input_indices:
                        break

        # Debug: show column matching results
        for col in INPUT_COLUMNS:
            if col in input_indices:
                idx = input_indices[col]
                raw = raw_headers[idx] if idx < len(raw_headers) else "?"
                if str(raw).strip() != col:
                    print(f"  Column: '{raw}' -> {col}")
            else:
                print(f"  WARNING: '{col}' not found. Headers: {headers}")

        if not input_indices:
            print(f"  WARNING: No matching columns found. Raw headers: {raw_headers}")

        parts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            part = {col: row[idx] for col, idx in input_indices.items()}
            parts.append(part)
        return parts

    def write_class_files(self, results: list[dict]) -> list[str]:
        """
        Group results by part_class and write one Excel per class.
        Each result dict must have keys: part (input dict), part_class, attributes, source_url.
        Returns list of output file paths written.
        """
        # Group by class
        by_class: dict[str, list[dict]] = {}
        for r in results:
            cls = r.get("part_class") or "Unclassified"
            by_class.setdefault(cls, []).append(r)

        written = []
        for cls, class_results in by_class.items():
            path = self._write_one_class(cls, class_results)
            written.append(path)
        return written

    # ── Private ───────────────────────────────────────────────────────────────

    def _write_one_class(self, cls: str, results: list[dict]) -> str:
        """Write a single per-class Excel file. Returns the file path.

        Column order: INPUT + PREFIX + METRICS + TC_ATTRS (navy) + LOV_MISMATCH (red) + AGENT_ATTRS (gray)
        TC attrs include inherited attributes from parent classes in Classes.json.
        """
        from src.attr_schema import get_schema

        # Get TC-configured attributes for this class (includes inherited from parents)
        tc_attrs = get_schema(cls)
        tc_attr_set = set(tc_attrs)

        # Collect all unique attribute keys across every part in this class
        all_attr_keys: list[str] = []
        seen: set[str] = set()
        for r in results:
            for k in r.get("attributes", {}).keys():
                if k not in seen:
                    all_attr_keys.append(k)
                    seen.add(k)

        # Split into TC attrs (ordered by schema) and agent-extracted attrs
        agent_attrs = [k for k in all_attr_keys if k not in tc_attr_set]

        # Build LOV mismatch columns: any attr that had a mismatch in any row
        lov_mismatch_attrs: list[str] = []
        seen_mm: set[str] = set()
        for r in results:
            for k in r.get("lov_mismatches", {}).keys():
                if k not in seen_mm and k in tc_attr_set:  # Only TC attrs for this class
                    lov_mismatch_attrs.append(k)
                    seen_mm.add(k)
        # Sort by TC schema position (O(1) per lookup); non-schema attrs go last
        tc_index = {name: i for i, name in enumerate(tc_attrs)}
        lov_mismatch_attrs.sort(key=lambda k: tc_index.get(k, 10**9))
        lov_mismatch_cols = [f"{k}-NotPresentInLOV" for k in lov_mismatch_attrs]

        # Build range-original columns: TC schema attrs where range was averaged in any row
        range_orig_attrs: list[str] = []
        seen_ro: set[str] = set()
        for r in results:
            for k in r.get("range_originals", {}).keys():
                if k not in seen_ro and k in tc_attr_set:
                    range_orig_attrs.append(k)
                    seen_ro.add(k)
        range_orig_attrs.sort(key=lambda k: tc_index.get(k, 10**9))
        range_orig_cols = [f"{k}-Original-RANGE-Value" for k in range_orig_attrs]

        # Build fraction-original columns: TC schema attrs where fractions were converted in any row
        fraction_orig_attrs: list[str] = []
        seen_fo: set[str] = set()
        for r in results:
            for k in r.get("fraction_originals", {}).keys():
                if k not in seen_fo and k in tc_attr_set:
                    fraction_orig_attrs.append(k)
                    seen_fo.add(k)
        fraction_orig_attrs.sort(key=lambda k: tc_index.get(k, 10**9))
        fraction_orig_cols = [f"{k}-Original-FRACTION-Value" for k in fraction_orig_attrs]

        # Build full column list: input + prefix + metrics + TC attrs + LOV-mismatch + range-orig + fraction-orig + agent attrs
        all_columns = (INPUT_COLUMNS + RESULT_PREFIX_COLS + METRICS_COLS
                       + tc_attrs + lov_mismatch_cols + range_orig_cols + fraction_orig_cols + agent_attrs)
        metrics_start = len(INPUT_COLUMNS) + len(RESULT_PREFIX_COLS)
        tc_attr_start = metrics_start + len(METRICS_COLS)
        lov_mismatch_start = tc_attr_start + len(tc_attrs)
        range_orig_start = lov_mismatch_start + len(lov_mismatch_cols)
        fraction_orig_start = range_orig_start + len(range_orig_cols)
        agent_attr_start = fraction_orig_start + len(fraction_orig_cols)

        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel sheet names can't contain / \ * ? : [ ]
        safe_title = cls.translate(str.maketrans("/\\*?:[]", "-------"))
        ws.title = safe_title[:31]

        # Write headers with color coding
        for col_idx, header in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Color: metrics (green), TC attrs (navy), LOV mismatch (red), range orig (steel blue), agent attrs (gray-purple)
            if metrics_start < col_idx <= tc_attr_start:
                cell.fill = METRICS_HEADER_FILL
            elif tc_attr_start < col_idx <= lov_mismatch_start:
                cell.fill = HEADER_FILL  # TC attrs (navy)
            elif lov_mismatch_start < col_idx <= range_orig_start:
                cell.fill = LOV_MISMATCH_HEADER_FILL  # LOV mismatch flag (red)
            elif range_orig_start < col_idx <= fraction_orig_start:
                cell.fill = RANGE_ORIG_HEADER_FILL  # original range values (steel blue)
            elif fraction_orig_start < col_idx <= agent_attr_start:
                cell.fill = FRACTION_ORIG_HEADER_FILL  # original fraction values (amber)
            elif col_idx > agent_attr_start:
                cell.fill = AGENT_HEADER_FILL  # agent extras (gray-purple)
            else:
                cell.fill = HEADER_FILL  # input + prefix (navy)
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        # Write data rows
        ordered_attr_keys = tc_attrs + agent_attrs  # combined in correct order
        for row_idx, r in enumerate(results, start=2):
            part = r.get("part", {})
            attrs = r.get("attributes", {})
            row_mismatches = r.get("lov_mismatches", {})

            # Input columns
            for col_idx, col_name in enumerate(INPUT_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=part.get(col_name))

            # Result prefix columns: Part Class, TC Class ID, Source URL, Unit of Measure (Agent)
            offset = len(INPUT_COLUMNS) + 1
            ws.cell(row=row_idx, column=offset, value=r.get("part_class", ""))
            ws.cell(row=row_idx, column=offset + 1, value=r.get("tc_class_id", ""))
            ws.cell(row=row_idx, column=offset + 2, value=str(r.get("source_url", "")))
            ws.cell(row=row_idx, column=offset + 3, value=r.get("unit_of_measure", ""))

            # Quality metrics columns
            _metrics_map = {
                "Extraction Coverage %": r.get("extraction_coverage", ""),
                "Source Reliability %": r.get("source_reliability", ""),
                "Classification Confidence %": r.get("classification_confidence", ""),
                "Source Type": r.get("source_type", ""),
                "LOV Compliance %": r.get("lov_compliance", ""),
                "Validation Action": r.get("validation_action", ""),
            }
            for mi, mc in enumerate(METRICS_COLS):
                val = _metrics_map.get(mc, "")
                cell = ws.cell(row=row_idx, column=offset + len(RESULT_PREFIX_COLS) + mi, value=val)
                # Color-code percentage cells
                if mc in _PCT_METRICS and isinstance(val, (int, float)):
                    if val >= 80:
                        cell.fill = GREEN_FILL
                    elif val >= 50:
                        cell.fill = AMBER_FILL
                    else:
                        cell.fill = RED_FILL
                    cell.font = Font(bold=True, color="FFFFFF")

            # TC attribute columns (no agent attrs yet)
            attr_offset = offset + len(RESULT_PREFIX_COLS) + len(METRICS_COLS)
            for i, key in enumerate(tc_attrs):
                ws.cell(row=row_idx, column=attr_offset + i, value=attrs.get(key, ""))

            # LOV mismatch columns (red headers, dark red cells with original value)
            mm_offset = attr_offset + len(tc_attrs)
            for i, mm_attr in enumerate(lov_mismatch_attrs):
                if mm_attr in row_mismatches:
                    cell = ws.cell(row=row_idx, column=mm_offset + i,
                                   value=row_mismatches[mm_attr])
                    cell.fill = LOV_MISMATCH_CELL_FILL
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    ws.cell(row=row_idx, column=mm_offset + i, value="")

            # Range original value columns (steel blue — original range before averaging)
            ro_offset = mm_offset + len(lov_mismatch_cols)
            row_range_originals = r.get("range_originals", {})
            for i, ro_attr in enumerate(range_orig_attrs):
                ws.cell(row=row_idx, column=ro_offset + i,
                        value=row_range_originals.get(ro_attr, ""))

            # Fraction original value columns (amber — original fraction before decimal conversion)
            fo_offset = ro_offset + len(range_orig_cols)
            row_fraction_originals = r.get("fraction_originals", {})
            for i, fo_attr in enumerate(fraction_orig_attrs):
                ws.cell(row=row_idx, column=fo_offset + i,
                        value=row_fraction_originals.get(fo_attr, ""))

            # Agent attribute columns (gray-purple)
            agent_offset = fo_offset + len(fraction_orig_cols)
            for i, key in enumerate(agent_attrs):
                ws.cell(row=row_idx, column=agent_offset + i, value=attrs.get(key, ""))

        # Auto-size columns
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        # Filename: add -NotInTc suffix if class not found in Classes.json
        import re
        safe_cls = re.sub(r'[\\/:*?"<>|]', "_", cls)
        # Check if any result has in_json=False (class not in TC)
        in_json = any(r.get("in_json", True) for r in results)
        suffix = "" if in_json else "-NotInTc"
        out_path = self.output_dir / f"{safe_cls}{suffix}.xlsx"
        wb.save(str(out_path))
        return str(out_path)
